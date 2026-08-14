import csv
import io
from datetime import date
from decimal import Decimal
from typing import List, Optional
from uuid import UUID

from sqlalchemy.ext.asyncio import AsyncSession

from app.models.transaction import Transaction
from app.repositories.transaction import TransactionRepository
from app.schemas.transaction import TransactionFilter


class ExportService:
    def __init__(self, db: AsyncSession):
        self.txn_repo = TransactionRepository(db)

    async def get_transactions(self, user_id: UUID, filters: TransactionFilter) -> List[Transaction]:
        return await self.txn_repo.get_all_for_export(user_id, filters)

    def to_csv(self, transactions: List[Transaction]) -> bytes:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "Date", "Type", "Category", "Account",
            "Amount", "Currency", "Description", "Notes", "Reference"
        ])
        for txn in transactions:
            writer.writerow([
                str(txn.transaction_date),
                txn.transaction_type.value,
                txn.category.name if txn.category else "",
                txn.account.name if txn.account else "",
                str(txn.amount),
                txn.currency,
                txn.description or "",
                txn.notes or "",
                txn.reference_number or "",
            ])
        return output.getvalue().encode("utf-8")

    def to_excel(self, transactions: List[Transaction]) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        headers = ["Date", "Type", "Category", "Account", "Amount", "Currency", "Description", "Notes", "Reference"]
        header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row, txn in enumerate(transactions, 2):
            ws.cell(row=row, column=1, value=str(txn.transaction_date))
            ws.cell(row=row, column=2, value=txn.transaction_type.value)
            ws.cell(row=row, column=3, value=txn.category.name if txn.category else "")
            ws.cell(row=row, column=4, value=txn.account.name if txn.account else "")
            ws.cell(row=row, column=5, value=float(txn.amount))
            ws.cell(row=row, column=6, value=txn.currency)
            ws.cell(row=row, column=7, value=txn.description or "")
            ws.cell(row=row, column=8, value=txn.notes or "")
            ws.cell(row=row, column=9, value=txn.reference_number or "")

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def to_pdf(self, transactions: List[Transaction]) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()

        elements.append(Paragraph("Transaction Report", styles["Title"]))
        elements.append(Spacer(1, 12))

        headers = ["Date", "Type", "Category", "Account", "Amount", "Description"]
        data = [headers]
        for txn in transactions:
            data.append([
                str(txn.transaction_date),
                txn.transaction_type.value,
                txn.category.name if txn.category else "",
                txn.account.name if txn.account else "",
                f"{txn.currency} {float(txn.amount):,.2f}",
                (txn.description or "")[:40],
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#166534")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0fdf4")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
        ]))
        elements.append(table)
        doc.build(elements)
        return output.getvalue()
